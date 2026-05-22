import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from server.excel_engine import (
    apply_print_plan,
    build_print_preview,
    build_workbook_preview,
    export_print_pdf,
    parse_a1_range,
    recommend_print_plans,
)


class ExcelEngineTests(unittest.TestCase):
    def make_workbook(self, root: Path) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sales"
        sheet.append(["Order", "Customer", "Region", "Amount", "Owner", "Status"])
        for index in range(1, 45):
            sheet.append([index, f"Customer {index}", "East", index * 12.5, "Lee", "Done"])
        notes = workbook.create_sheet("Notes")
        notes.append(["Info", "Value"])
        notes.append(["Quarter", "Q2"])
        path = root / "sales.xlsx"
        workbook.save(path)
        return path

    def test_preview_reads_real_sheet_names_ranges_and_cells(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = self.make_workbook(Path(temp_dir))

            preview = build_workbook_preview(source)

            self.assertEqual(preview["filename"], "sales.xlsx")
            self.assertEqual([sheet["name"] for sheet in preview["sheets"]], ["Sales", "Notes"])
            self.assertEqual(preview["sheets"][0]["range"], "A1:F45")
            self.assertEqual(preview["sheets"][0]["autoRange"], "A1:F45")
            self.assertEqual(preview["sheets"][0]["printRange"], "A1:F45")
            self.assertEqual(preview["sheets"][0]["preview"][0][0], "Order")

    def test_preview_uses_existing_excel_print_area_when_present(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = self.make_workbook(root)
            workbook = load_workbook(source)
            sheet = workbook["Sales"]
            sheet.print_area = "B2:D8"
            workbook.save(source)

            preview = build_workbook_preview(source)

            self.assertTrue(preview["sheets"][0]["hasNativePrintArea"])
            self.assertEqual(preview["sheets"][0]["nativePrintArea"], "B2:D8")
            self.assertEqual(preview["sheets"][0]["printRange"], "B2:D8")
            self.assertEqual(preview["sheets"][0]["rangeSource"], "excel")
            self.assertEqual(preview["sheets"][0]["preview"][0][0], "Customer 1")

    def test_preview_preserves_numeric_display_format(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = self.make_workbook(root)
            workbook = load_workbook(source)
            sheet = workbook["Sales"]
            sheet["D2"] = 1234.5
            sheet["D2"].number_format = '#,##0.00'
            workbook.save(source)

            workbook_preview = build_workbook_preview(source)
            print_preview = build_print_preview(
                source,
                [{"sheet": "Sales", "range": "A1:F45", "rangeSource": "manual"}],
                {
                    "id": "fit-columns",
                    "orientation": "landscape",
                    "fitMode": "fitColumns",
                    "fitToWidth": 1,
                    "fitToHeight": 0,
                    "repeatHeader": True,
                },
            )

            self.assertEqual(workbook_preview["sheets"][0]["preview"][1][3], "1,234.50")
            self.assertEqual(print_preview["sheets"][0]["pages"][0]["rows"][1][3], "1,234.50")

    def test_recommended_plans_include_print_ready_choices(self):
        plans = recommend_print_plans(
            {
                "name": "Sales",
                "rows": 45,
                "columns": 6,
                "range": "A1:F45",
            }
        )

        self.assertGreaterEqual(len(plans), 5)
        self.assertEqual(plans[0]["id"], "fit-columns")

    def test_apply_plan_sets_print_area_titles_orientation_and_outputs_xlsx(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = self.make_workbook(root)
            output = root / "optimized.xlsx"

            apply_print_plan(
                source,
                output,
                [{"sheet": "Sales", "range": "A1:F45", "rangeSource": "manual"}],
                {
                    "id": "fit-columns",
                    "orientation": "landscape",
                    "fitMode": "fitColumns",
                    "fitToWidth": 1,
                    "fitToHeight": 0,
                    "repeatHeader": True,
                    "freezeTopRow": True,
                },
            )

            workbook = load_workbook(output)
            sheet = workbook["Sales"]
            self.assertEqual(sheet.print_area, "'Sales'!$A$1:$F$45")
            self.assertEqual(sheet.print_title_rows, "$1:$1")
            self.assertEqual(sheet.page_setup.orientation, "landscape")
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertEqual([page_break.id for page_break in sheet.row_breaks.brk], [27])

    def test_print_preview_simulates_pages_before_generation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = self.make_workbook(Path(temp_dir))

            preview = build_print_preview(
                source,
                [{"sheet": "Sales", "range": "A1:F45", "rangeSource": "manual"}],
                {
                    "id": "fit-columns",
                    "orientation": "landscape",
                    "fitMode": "fitColumns",
                    "fitToWidth": 1,
                    "fitToHeight": 0,
                    "repeatHeader": True,
                },
            )

            self.assertEqual(preview["selectedSheetCount"], 1)
            self.assertEqual(preview["optimizedPageCount"], 2)
            self.assertGreater(preview["originalPageCount"], preview["optimizedPageCount"])
            self.assertEqual(preview["sheets"][0]["pages"][0]["rowRange"], "1-26")
            self.assertEqual(preview["sheets"][0]["pages"][1]["rows"][0][0], "Order")
            self.assertEqual(preview["sheets"][0]["pages"][1]["rows"][1][0], 26)

    def test_single_page_preview_keeps_full_page_content(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = self.make_workbook(Path(temp_dir))

            preview = build_print_preview(
                source,
                [{"sheet": "Sales", "range": "A1:F45", "rangeSource": "manual"}],
                {
                    "id": "single-page",
                    "orientation": "portrait",
                    "fitMode": "singlePage",
                    "fitToWidth": 1,
                    "fitToHeight": 1,
                    "repeatHeader": True,
                },
            )

            self.assertEqual(preview["sheets"][0]["optimizedPageCount"], 1)
            self.assertEqual(len(preview["sheets"][0]["pages"][0]["rows"]), 45)
            self.assertEqual(preview["sheets"][0]["pages"][0]["rows"][-1][0], 44)

    def test_preview_respects_manual_print_range(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = self.make_workbook(Path(temp_dir))

            preview = build_print_preview(
                source,
                [{"sheet": "Sales", "range": "B2:D8", "rangeSource": "manual"}],
                {
                    "id": "fit-rows",
                    "orientation": "portrait",
                    "fitMode": "fitRows",
                    "fitToWidth": 0,
                    "fitToHeight": 1,
                    "repeatHeader": True,
                },
            )

            self.assertEqual(preview["sheets"][0]["range"], "B2:D8")
            self.assertEqual(preview["sheets"][0]["pages"][0]["columnRange"], "B-D")
            self.assertEqual(preview["sheets"][0]["pages"][0]["rows"][0][0], "Customer 1")

    def test_preview_and_output_can_use_per_sheet_print_settings(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = self.make_workbook(root)
            output = root / "per-sheet.xlsx"

            targets = [
                {
                    "sheet": "Sales",
                    "range": "A1:F45",
                    "rangeSource": "manual",
                    "plan": {
                        "id": "fit-columns",
                        "orientation": "landscape",
                        "fitMode": "fitColumns",
                        "fitToWidth": 1,
                        "fitToHeight": 0,
                        "repeatHeader": True,
                        "freezeTopRow": True,
                    },
                },
                {
                    "sheet": "Notes",
                    "range": "A1:B2",
                    "rangeSource": "manual",
                    "plan": {
                        "id": "single-page",
                        "orientation": "portrait",
                        "fitMode": "singlePage",
                        "fitToWidth": 1,
                        "fitToHeight": 1,
                        "repeatHeader": False,
                        "freezeTopRow": False,
                    },
                },
            ]

            preview = build_print_preview(source, targets, {"id": "fit-columns"})
            apply_print_plan(source, output, targets, {"id": "fit-columns"})

            sales_preview = next(sheet for sheet in preview["sheets"] if sheet["name"] == "Sales")
            notes_preview = next(sheet for sheet in preview["sheets"] if sheet["name"] == "Notes")
            workbook = load_workbook(output)

            self.assertTrue(preview["hasMixedPlans"])
            self.assertEqual(sales_preview["plan"]["orientation"], "landscape")
            self.assertEqual(notes_preview["plan"]["orientation"], "portrait")
            self.assertEqual(workbook["Sales"].page_setup.orientation, "landscape")
            self.assertEqual(workbook["Notes"].page_setup.orientation, "portrait")
            self.assertEqual(workbook["Notes"].page_setup.fitToHeight, 1)

    def test_pdf_export_writes_printable_pdf_bytes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = self.make_workbook(root)
            output = root / "print.pdf"

            export_print_pdf(source, output, ["Sales"])

            self.assertTrue(output.exists())
            self.assertTrue(output.read_bytes().startswith(b"%PDF"))

    def test_parse_a1_range_accepts_single_cell(self):
        self.assertEqual(parse_a1_range("C7"), (7, 3, 7, 3))

    def test_apply_plan_can_wrap_selected_long_text_columns(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = self.make_workbook(root)
            workbook = load_workbook(source)
            sheet = workbook["Sales"]
            sheet["B2"] = "This is a very long customer name that should wrap for printing"
            workbook.save(source)

            output = root / "wrapped.xlsx"
            apply_print_plan(
                source,
                output,
                [{"sheet": "Sales", "range": "A1:F45", "rangeSource": "manual"}],
                {
                    "id": "fit-columns",
                    "orientation": "landscape",
                    "fitMode": "fitColumns",
                    "fitToWidth": 1,
                    "fitToHeight": 0,
                    "wrapLongText": True,
                    "wrapTextColumns": "B",
                },
            )

            wrapped = load_workbook(output)
            wrapped_sheet = wrapped["Sales"]
            self.assertTrue(wrapped_sheet["B2"].alignment.wrap_text)
            self.assertEqual(wrapped_sheet["B2"].alignment.vertical, "center")
            self.assertLessEqual(wrapped_sheet.column_dimensions["B"].width, 16)
            self.assertGreater(wrapped_sheet.row_dimensions[2].height, 18)


if __name__ == "__main__":
    unittest.main()
