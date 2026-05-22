from __future__ import annotations

import json
import math
import re
import subprocess
import tempfile
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.pagebreak import Break
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Table, TableStyle


PLAN_LIBRARY = [
    {
        "id": "fit-columns",
        "name": "标准归档版",
        "description": "适合常规明细表，按宽度压到一页并保留表头。",
        "orientation": "landscape",
        "fitMode": "fitColumns",
        "fitToWidth": 1,
        "fitToHeight": 0,
        "repeatHeader": True,
        "freezeTopRow": True,
        "margin": "normal",
        "beautify": True,
        "showGridLines": False,
        "wrapLongText": False,
        "wrapTextColumns": "",
    },
    {
        "id": "single-page",
        "name": "整表一页版",
        "description": "小型表格整页压缩，适合汇报或单页归档。",
        "orientation": "portrait",
        "fitMode": "singlePage",
        "fitToWidth": 1,
        "fitToHeight": 1,
        "repeatHeader": True,
        "freezeTopRow": True,
        "margin": "normal",
        "beautify": True,
        "showGridLines": False,
        "wrapLongText": False,
        "wrapTextColumns": "",
    },
    {
        "id": "fit-rows",
        "name": "长表纵向版",
        "description": "适合窄而长的表格，尽量把行压缩在一页高度内。",
        "orientation": "portrait",
        "fitMode": "fitRows",
        "fitToWidth": 0,
        "fitToHeight": 1,
        "repeatHeader": True,
        "freezeTopRow": True,
        "margin": "normal",
        "beautify": True,
        "showGridLines": False,
        "wrapLongText": False,
        "wrapTextColumns": "",
    },
    {
        "id": "compact",
        "name": "节纸紧凑版",
        "description": "边距更窄，适合批量归档和节省纸张。",
        "orientation": "landscape",
        "fitMode": "fitColumns",
        "fitToWidth": 1,
        "fitToHeight": 0,
        "repeatHeader": True,
        "freezeTopRow": True,
        "margin": "compact",
        "beautify": True,
        "showGridLines": False,
        "wrapLongText": False,
        "wrapTextColumns": "",
    },
    {
        "id": "manual",
        "name": "手动控制版",
        "description": "保留高级设置入口，按当前页面设置输出。",
        "orientation": "landscape",
        "fitMode": "fitColumns",
        "fitToWidth": 1,
        "fitToHeight": 0,
        "repeatHeader": False,
        "freezeTopRow": False,
        "margin": "normal",
        "beautify": False,
        "showGridLines": True,
        "wrapLongText": False,
        "wrapTextColumns": "",
    },
]

CELL_RANGE_RE = re.compile(
    r"^\$?([A-Za-z]{1,3})\$?(\d+)(?::\$?([A-Za-z]{1,3})\$?(\d+))?$"
)


def json_safe(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def column_index(label: str) -> int:
    total = 0
    for char in label.upper():
        total = total * 26 + (ord(char) - 64)
    return total


def parse_a1_range(value: str) -> tuple[int, int, int, int]:
    match = CELL_RANGE_RE.match((value or "").strip())
    if not match:
        raise ValueError(f"Invalid A1 range: {value}")
    start_col = column_index(match.group(1))
    start_row = int(match.group(2))
    end_col = column_index(match.group(3) or match.group(1))
    end_row = int(match.group(4) or match.group(2))
    min_row = min(start_row, end_row)
    max_row = max(start_row, end_row)
    min_col = min(start_col, end_col)
    max_col = max(start_col, end_col)
    return min_row, min_col, max_row, max_col


def used_range(sheet) -> tuple[int, int, int, int]:
    rows: list[int] = []
    columns: list[int] = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                rows.append(cell.row)
                columns.append(cell.column)
    if not rows or not columns:
        return 1, 1, 1, 1
    return min(rows), min(columns), max(rows), max(columns)


def range_label(bounds: tuple[int, int, int, int]) -> str:
    min_row, min_col, max_row, max_col = bounds
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def native_print_area(sheet) -> tuple[str | None, tuple[int, int, int, int] | None]:
    raw_value = getattr(sheet, "print_area", None)
    if not raw_value:
        return None, None
    if isinstance(raw_value, (list, tuple)):
        candidates = [str(item or "").strip() for item in raw_value]
    else:
        candidates = [item.strip() for item in str(raw_value).split(",")]
    for candidate in candidates:
        if not candidate:
            continue
        area = candidate.split("!", 1)[-1].replace("$", "").strip()
        if ":" not in area and not CELL_RANGE_RE.match(area):
            continue
        try:
            bounds = parse_a1_range(area)
        except ValueError:
            continue
        return range_label(bounds), bounds
    return None, None


def parse_column_spec(spec: str, min_col: int, max_col: int) -> list[int]:
    if not spec:
        return []
    selected: set[int] = set()
    for raw_part in str(spec).split(","):
        part = raw_part.strip().replace("$", "")
        if not part:
            continue
        if ":" in part:
            start, end = [item.strip() for item in part.split(":", 1)]
            if not start or not end:
                continue
            start_col = column_index(start)
            end_col = column_index(end)
            for col in range(min(start_col, end_col), max(start_col, end_col) + 1):
                if min_col <= col <= max_col:
                    selected.add(col)
        else:
            col = column_index(part)
            if min_col <= col <= max_col:
                selected.add(col)
    return sorted(selected)


def intersects(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> bool:
    return not (
        a[2] < b[0]
        or b[2] < a[0]
        or a[3] < b[1]
        or b[3] < a[1]
    )


def contains(outer: tuple[int, int, int, int], inner: tuple[int, int, int, int]) -> bool:
    return (
        outer[0] <= inner[0]
        and outer[1] <= inner[1]
        and outer[2] >= inner[2]
        and outer[3] >= inner[3]
    )


def merged_regions(sheet) -> list[tuple[int, int, int, int]]:
    return [
        (item.min_row, item.min_col, item.max_row, item.max_col)
        for item in sheet.merged_cells.ranges
    ]


def merged_regions_in_bounds(sheet, bounds: tuple[int, int, int, int]) -> list[dict[str, Any]]:
    items = []
    for min_row, min_col, max_row, max_col in merged_regions(sheet):
        region = (min_row, min_col, max_row, max_col)
        if intersects(region, bounds):
            items.append(
                {
                    "startRow": min_row,
                    "startCol": min_col,
                    "endRow": max_row,
                    "endCol": max_col,
                    "label": range_label(region),
                }
            )
    return items


def expand_bounds_for_merges(sheet, bounds: tuple[int, int, int, int]) -> tuple[tuple[int, int, int, int], list[str]]:
    expanded = bounds
    adjustments: list[str] = []
    changed = True
    while changed:
        changed = False
        for region in merged_regions(sheet):
            if intersects(region, expanded) and not contains(expanded, region):
                expanded = (
                    min(expanded[0], region[0]),
                    min(expanded[1], region[1]),
                    max(expanded[2], region[2]),
                    max(expanded[3], region[3]),
                )
                label = range_label(region)
                if label not in adjustments:
                    adjustments.append(label)
                changed = True
    return expanded, adjustments


def clamp_grid_bounds(bounds: tuple[int, int, int, int], max_rows: int = 80, max_columns: int = 18) -> tuple[int, int, int, int]:
    min_row, min_col, max_row, max_col = bounds
    return (
        min_row,
        min_col,
        min(max_row, min_row + max_rows - 1),
        min(max_col, min_col + max_columns - 1),
    )


def display_value(formula_cell, value_cell) -> tuple[Any, bool]:
    formula = formula_cell.value
    if isinstance(formula, str) and formula.startswith("="):
        if value_cell.value is None:
            return "公式待重新计算", True
        return value_cell.value, False
    return formula, False


def cleaned_number_format(number_format: str) -> str:
    text = str(number_format or "").strip()
    if not text:
        return ""
    text = re.sub(r"\[\$([^\]-]+)[^\]]*\]", r"\1", text)
    text = re.sub(r"\[[^\]]+\]", "", text)
    text = re.sub(r"_(.)", "", text)
    text = re.sub(r"\*(.)", "", text)
    return text.replace("\\", "")


def format_literal(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r'"([^"]*)"', r"\1", text)
    return text.replace("?", " ").strip()


def choose_format_section(number_format: str, value: int | float) -> str:
    parts = cleaned_number_format(number_format).split(";")
    if value > 0 or len(parts) == 1:
        return parts[0]
    if value < 0 and len(parts) > 1:
        return parts[1]
    if value == 0 and len(parts) > 2:
        return parts[2]
    return parts[0]


def format_numeric_value(value: int | float, number_format: str) -> Any:
    section = choose_format_section(number_format, value)
    if not section or section.lower() == "general" or "@" in section or is_date_format(number_format):
        return value

    placeholder_indexes = [index for index, char in enumerate(section) if char in "0#?.,%"]
    if not placeholder_indexes:
        return value

    first = placeholder_indexes[0]
    last = placeholder_indexes[-1]
    prefix = format_literal(section[:first])
    suffix = format_literal(section[last + 1 :])
    core = section[first : last + 1]
    working_value = Decimal(str(abs(value) if value < 0 and ";" in cleaned_number_format(number_format) else value))

    percent_count = core.count("%")
    if percent_count:
        working_value *= Decimal(100) ** percent_count

    decimal_part = core.split(".", 1)[1] if "." in core else ""
    decimals = sum(1 for char in decimal_part if char in "0#?")
    quantizer = Decimal("1") if decimals == 0 else Decimal(f"1.{'0' * decimals}")
    rounded = working_value.quantize(quantizer, rounding=ROUND_HALF_UP)

    integer_part = core.split(".", 1)[0]
    fmt = f",.{decimals}f" if "," in integer_part else f".{decimals}f"
    number_text = format(rounded, fmt)
    if decimals == 0 and "." in number_text:
        number_text = number_text.split(".", 1)[0]
    if value < 0 and ";" not in cleaned_number_format(number_format) and not number_text.startswith("-"):
        number_text = f"-{number_text}"
    return f"{prefix}{number_text}{suffix}"


def display_output(formula_cell, value_cell) -> tuple[Any, bool]:
    value, missing_cached = display_value(formula_cell, value_cell)
    if isinstance(value, bool):
        return value, missing_cached
    if isinstance(value, (int, float)):
        return format_numeric_value(value, formula_cell.number_format), missing_cached
    return value, missing_cached


def sheet_grid(formula_sheet, value_sheet, bounds: tuple[int, int, int, int], max_rows: int = 80, max_columns: int = 18) -> dict[str, Any]:
    grid_bounds = clamp_grid_bounds(bounds, max_rows=max_rows, max_columns=max_columns)
    min_row, min_col, max_row, max_col = grid_bounds
    merge_map: dict[tuple[int, int], dict[str, Any]] = {}
    for region in merged_regions_in_bounds(formula_sheet, grid_bounds):
        for row in range(region["startRow"], region["endRow"] + 1):
            for col in range(region["startCol"], region["endCol"] + 1):
                merge_map[(row, col)] = region

    warnings: list[dict[str, Any]] = []
    cells: list[list[dict[str, Any]]] = []
    for row in range(min_row, max_row + 1):
        current_row = []
        for col in range(min_col, max_col + 1):
            formula_cell = formula_sheet.cell(row=row, column=col)
            value_cell = value_sheet.cell(row=row, column=col)
            region = merge_map.get((row, col))
            is_anchor = not region or (row == region["startRow"] and col == region["startCol"])
            value, missing_cached = display_output(formula_cell, value_cell)
            if missing_cached:
                warnings.append(
                    {
                        "cell": f"{get_column_letter(col)}{row}",
                        "message": "该公式单元格没有可用缓存值，建议在 Excel/WPS 中重新计算并保存后再上传。",
                    }
                )
            current_row.append(
                {
                    "row": row,
                    "col": col,
                    "ref": f"{get_column_letter(col)}{row}",
                    "display": json_safe(value),
                    "formula": formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else "",
                    "missingCachedValue": missing_cached,
                    "isMerged": bool(region),
                    "isMergeAnchor": is_anchor,
                    "rowSpan": (region["endRow"] - region["startRow"] + 1) if region and is_anchor else 1,
                    "colSpan": (region["endCol"] - region["startCol"] + 1) if region and is_anchor else 1,
                    "hiddenByMerge": bool(region and not is_anchor),
                }
            )
        cells.append(current_row)

    return {
        "startRow": min_row,
        "endRow": max_row,
        "startCol": min_col,
        "endCol": max_col,
        "cells": cells,
        "mergedRegions": merged_regions_in_bounds(formula_sheet, grid_bounds),
        "formulaWarnings": warnings,
        "truncated": grid_bounds != bounds,
    }


def analyze_sheet(formula_sheet, value_sheet, include_grid: bool = True) -> dict[str, Any]:
    bounds = used_range(formula_sheet)
    native_area_label, native_bounds = native_print_area(formula_sheet)
    min_row, min_col, max_row, max_col = bounds
    row_count = max_row - min_row + 1
    column_count = max_col - min_col + 1
    active_bounds = native_bounds or bounds
    active_range = native_area_label or range_label(bounds)
    compact_grid = sheet_grid(formula_sheet, value_sheet, active_bounds, max_rows=12, max_columns=12)
    full_grid = sheet_grid(formula_sheet, value_sheet, active_bounds) if include_grid else None
    issues = []
    if column_count > 8:
        issues.append({"id": "wide", "label": "列较多，建议按列压缩或横向打印"})
    if row_count > 35:
        issues.append({"id": "header", "label": "行数较多，建议重复打印表头"})
    if formula_sheet.merged_cells.ranges:
        issues.append({"id": "merged", "label": "包含合并单元格，打印区域会自动保护合并块"})
    if native_area_label:
        issues.insert(0, {"id": "native-print-area", "label": f"已检测到 Excel 原打印区域 {native_area_label}"})
    if any(
        item["missingCachedValue"]
        for row in compact_grid["cells"]
        for item in row
    ):
        issues.append({"id": "formula", "label": "存在公式缓存缺失，预览可能与 Excel 重新计算结果不同"})
    return {
        "name": formula_sheet.title,
        "range": range_label(bounds),
        "autoRange": range_label(bounds),
        "printRange": active_range,
        "rangeSource": "excel" if native_area_label else "auto",
        "hasNativePrintArea": bool(native_area_label),
        "nativePrintArea": native_area_label or "",
        "rows": row_count,
        "columns": column_count,
        "preview": [
            [cell["display"] for cell in row if not cell["hiddenByMerge"]]
            for row in compact_grid["cells"][:8]
        ],
        "grid": full_grid,
        "issues": issues,
        "recommendedOrientation": "landscape" if column_count > 8 else "portrait",
        "hidden": formula_sheet.sheet_state != "visible",
        "mergedRegions": merged_regions_in_bounds(formula_sheet, active_bounds),
    }


def build_workbook_preview(source: str | Path, include_grids: bool = True) -> dict[str, Any]:
    source_path = Path(source)
    workbook_formula = load_workbook(source_path, data_only=False)
    workbook_values = load_workbook(source_path, data_only=True)
    sheets = []
    for formula_sheet, value_sheet in zip(workbook_formula.worksheets, workbook_values.worksheets):
        if formula_sheet.sheet_state != "visible":
            continue
        sheets.append(analyze_sheet(formula_sheet, value_sheet, include_grid=include_grids))
    return {
        "filename": source_path.name,
        "sheetCount": len(sheets),
        "sheets": sheets,
        "plans": recommend_print_plans(sheets[0] if sheets else {}),
    }


def build_sheet_detail(source: str | Path, sheet_name: str) -> dict[str, Any]:
    source_path = Path(source)
    workbook_formula = load_workbook(source_path, data_only=False)
    workbook_values = load_workbook(source_path, data_only=True)
    for formula_sheet, value_sheet in zip(workbook_formula.worksheets, workbook_values.worksheets):
        if formula_sheet.title != sheet_name:
            continue
        if formula_sheet.sheet_state != "visible":
            raise ValueError(f"{sheet_name} 不是可见 Sheet")
        return analyze_sheet(formula_sheet, value_sheet, include_grid=True)
    raise ValueError(f"未找到 Sheet：{sheet_name}")


def recommend_print_plans(sheet: dict[str, Any]) -> list[dict[str, Any]]:
    plans = [dict(plan) for plan in PLAN_LIBRARY]
    columns = sheet.get("columns", 0)
    rows = sheet.get("rows", 0)
    if columns > 10:
        first = next(index for index, item in enumerate(plans) if item["id"] == "fit-columns")
        plans.insert(0, plans.pop(first))
    elif rows < 20:
        first = next(index for index, item in enumerate(plans) if item["id"] == "single-page")
        plans.insert(0, plans.pop(first))
    return plans


def normalize_fit_mode(config: dict[str, Any]) -> dict[str, Any]:
    fit_mode = config.get("fitMode")
    if fit_mode == "singlePage":
        config["fitToWidth"] = 1
        config["fitToHeight"] = 1
    elif fit_mode == "fitRows":
        config["fitToWidth"] = 0
        config["fitToHeight"] = 1
    elif fit_mode == "fitColumns":
        config["fitToWidth"] = 1
        config["fitToHeight"] = 0
    config.setdefault("fitMode", "fitColumns")
    config.setdefault("fitToWidth", 1)
    config.setdefault("fitToHeight", 0)
    config.setdefault("beautify", True)
    config.setdefault("showGridLines", False)
    config.setdefault("wrapLongText", False)
    config.setdefault("wrapTextColumns", "")
    return config


def resolve_plan(plan: str | dict[str, Any]) -> dict[str, Any]:
    if isinstance(plan, dict):
        base = next((item for item in PLAN_LIBRARY if item["id"] == plan.get("id")), PLAN_LIBRARY[0])
        return normalize_fit_mode({**base, **plan})
    return normalize_fit_mode(next((item for item in PLAN_LIBRARY if item["id"] == plan), PLAN_LIBRARY[0]).copy())


def normalize_print_targets(selected_targets: Iterable[Any]) -> list[dict[str, Any]]:
    normalized = []
    for item in selected_targets:
        if isinstance(item, str):
            normalized.append({"sheet": item, "rangeSource": "auto"})
            continue
        if not isinstance(item, dict):
            continue
        sheet_name = item.get("sheet") or item.get("name")
        if not sheet_name:
            continue
        normalized.append(
            {
                "sheet": sheet_name,
                "range": item.get("range"),
                "rangeSource": item.get("rangeSource", "auto"),
                "plan": item.get("plan"),
            }
        )
    return normalized


def target_plan_config(target: dict[str, Any], fallback_plan: str | dict[str, Any]) -> dict[str, Any]:
    return resolve_plan(target.get("plan") or fallback_plan)


def target_bounds(sheet, target: dict[str, Any]) -> tuple[tuple[int, int, int, int], list[str]]:
    auto_bounds = used_range(sheet)
    requested = target.get("range") or range_label(auto_bounds)
    bounds = parse_a1_range(requested)
    if not intersects(bounds, auto_bounds):
        raise ValueError(f"{sheet.title} 的打印区域超出了已识别的数据范围")
    bounds = (
        max(bounds[0], auto_bounds[0]),
        max(bounds[1], auto_bounds[1]),
        min(bounds[2], auto_bounds[2]),
        min(bounds[3], auto_bounds[3]),
    )
    return expand_bounds_for_merges(sheet, bounds)


def preview_rows_per_page(config: dict[str, Any]) -> int:
    rows = 26 if config.get("orientation") == "landscape" else 34
    if config.get("margin") == "compact":
        rows += 6
    return rows


def preview_columns_per_page(config: dict[str, Any]) -> int:
    columns = 8 if config.get("orientation") == "landscape" else 6
    if config.get("margin") == "compact":
        columns += 1
    return columns


def build_groups(start: int, end: int, fit: int, fallback_page_size: int) -> list[tuple[int, int]]:
    total = end - start + 1
    if fit == 1:
        return [(start, end)]
    if fit == 0:
        return [
            (group_start, min(end, group_start + fallback_page_size - 1))
            for group_start in range(start, end + 1, fallback_page_size)
        ]
    group_size = max(1, math.ceil(total / fit))
    return [
        (group_start, min(end, group_start + group_size - 1))
        for group_start in range(start, end + 1, group_size)
    ]


def page_cells(formula_sheet, value_sheet, row_start: int, row_end: int, col_start: int, col_end: int) -> list[list[Any]]:
    rows = []
    for row in range(row_start, row_end + 1):
        current = []
        for col in range(col_start, col_end + 1):
            formula_cell = formula_sheet.cell(row=row, column=col)
            value_cell = value_sheet.cell(row=row, column=col)
            display, _ = display_output(formula_cell, value_cell)
            current.append(json_safe(display))
        rows.append(current)
    return rows


def simulate_sheet_pages(formula_sheet, value_sheet, bounds: tuple[int, int, int, int], config: dict[str, Any], adjustments: list[str] | None = None) -> dict[str, Any]:
    min_row, min_col, max_row, max_col = bounds
    row_groups = build_groups(min_row, max_row, int(config.get("fitToHeight", 0)), preview_rows_per_page(config))
    column_groups = build_groups(min_col, max_col, int(config.get("fitToWidth", 1)), preview_columns_per_page(config))
    header = page_cells(formula_sheet, value_sheet, min_row, min_row, min_col, max_col)[0]
    pages = []
    page_number = 1

    for col_start, col_end in column_groups:
        for row_start, row_end in row_groups:
            rows = page_cells(formula_sheet, value_sheet, row_start, row_end, col_start, col_end)
            if config.get("repeatHeader") and row_start > min_row:
                rows = page_cells(formula_sheet, value_sheet, min_row, min_row, col_start, col_end) + rows
            pages.append(
                {
                    "number": page_number,
                    "rowRange": f"{row_start}-{row_end}",
                    "columnRange": f"{get_column_letter(col_start)}-{get_column_letter(col_end)}",
                    "rows": rows,
                    "hasRepeatedHeader": bool(config.get("repeatHeader") and row_start > min_row),
                }
            )
            page_number += 1

    raw_rows = max(1, math.ceil((max_row - min_row + 1) / 18))
    raw_cols = max(1, math.ceil((max_col - min_col + 1) / 8))
    return {
        "name": formula_sheet.title,
        "range": range_label(bounds),
        "plan": config,
        "rows": max_row - min_row + 1,
        "columns": max_col - min_col + 1,
        "header": header,
        "pages": pages,
        "optimizedPageCount": len(pages),
        "originalPageCount": raw_rows * raw_cols,
        "rangeAdjustedByMerges": adjustments or [],
        "formulaWarnings": sheet_grid(formula_sheet, value_sheet, bounds, max_rows=18, max_columns=18)["formulaWarnings"],
    }


def build_print_preview(source: str | Path, selected_sheet_names: Iterable[Any], plan: str | dict[str, Any]) -> dict[str, Any]:
    targets = normalize_print_targets(selected_sheet_names)
    selected_names = {target["sheet"] for target in targets}
    fallback_config = resolve_plan(plan)
    workbook_formula = load_workbook(source, data_only=False)
    workbook_values = load_workbook(source, data_only=True)
    target_map = {target["sheet"]: target for target in targets}
    sheets = []
    for formula_sheet, value_sheet in zip(workbook_formula.worksheets, workbook_values.worksheets):
        if formula_sheet.title not in selected_names or formula_sheet.sheet_state != "visible":
            continue
        target = target_map[formula_sheet.title]
        config = target_plan_config(target, fallback_config)
        bounds, adjustments = target_bounds(formula_sheet, target)
        sheets.append(simulate_sheet_pages(formula_sheet, value_sheet, bounds, config, adjustments))
    distinct_plan_count = len({json.dumps(sheet["plan"], ensure_ascii=False, sort_keys=True) for sheet in sheets}) if sheets else 0
    return {
        "plan": fallback_config,
        "hasMixedPlans": distinct_plan_count > 1,
        "selectedSheetCount": len(sheets),
        "originalPageCount": sum(sheet["originalPageCount"] for sheet in sheets),
        "optimizedPageCount": sum(sheet["optimizedPageCount"] for sheet in sheets),
        "sheets": sheets,
    }


def set_margins(sheet, margin: str) -> None:
    if margin == "compact":
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.35
        sheet.page_margins.bottom = 0.35
    else:
        sheet.page_margins.left = 0.4
        sheet.page_margins.right = 0.4
        sheet.page_margins.top = 0.55
        sheet.page_margins.bottom = 0.55


def beautify_range(sheet, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1B7F72")
    soft_fill = PatternFill(fill_type="solid", fgColor="F4FBF9")
    border = Border(
        left=Side(style="thin", color="D3DEE8"),
        right=Side(style="thin", color="D3DEE8"),
        top=Side(style="thin", color="D3DEE8"),
        bottom=Side(style="thin", color="D3DEE8"),
    )
    for col in range(min_col, max_col + 1):
        cell = sheet.cell(row=min_row, column=col)
        cell.font = Font(name=cell.font.name or "Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in range(min_row + 1, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border
            if (row - min_row) % 2 == 0:
                cell.fill = soft_fill


def apply_wrapped_columns(sheet, min_row: int, min_col: int, max_row: int, max_col: int, config: dict[str, Any]) -> None:
    if not config.get("wrapLongText"):
        return
    target_columns = parse_column_spec(str(config.get("wrapTextColumns") or ""), min_col, max_col)
    if not target_columns:
        longest = []
        for col in range(min_col, max_col + 1):
            values = [
                str(sheet.cell(row=row, column=col).value or "")
                for row in range(min_row + 1, min(max_row, min_row + 25) + 1)
            ]
            score = max((len(value) for value in values), default=0)
            if score >= 16:
                longest.append((score, col))
        longest.sort(reverse=True)
        target_columns = [col for _, col in longest[:2]]
    if not target_columns:
        return

    for col in target_columns:
        letter = get_column_letter(col)
        sheet.column_dimensions[letter].width = min(sheet.column_dimensions[letter].width or 16, 16)

    for row in range(min_row + 1, max_row + 1):
        max_lines = 1
        for col in target_columns:
            cell = sheet.cell(row=row, column=col)
            value = "" if cell.value is None else str(cell.value)
            if not value:
                continue
            approx_lines = max(1, math.ceil(len(value) / 16))
            max_lines = max(max_lines, approx_lines)
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "left",
                vertical="top",
                wrap_text=True,
            )
        if max_lines > 1:
            sheet.row_dimensions[row].height = max(18, 15 * max_lines)


def apply_print_plan(source: str | Path, output: str | Path, selected_sheet_names: Iterable[Any], plan: str | dict[str, Any]) -> Path:
    targets = normalize_print_targets(selected_sheet_names)
    selected = {target["sheet"] for target in targets}
    fallback_config = resolve_plan(plan)
    workbook = load_workbook(source)
    target_map = {target["sheet"]: target for target in targets}
    for sheet in workbook.worksheets:
        if sheet.title not in selected:
            continue
        target = target_map[sheet.title]
        config = target_plan_config(target, fallback_config)
        bounds, _ = target_bounds(sheet, target)
        min_row, min_col, max_row, max_col = bounds
        sheet.print_area = range_label(bounds)
        sheet.page_setup.orientation = config["orientation"]
        sheet.page_setup.fitToWidth = int(config.get("fitToWidth", 1))
        sheet.page_setup.fitToHeight = int(config.get("fitToHeight", 0))
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_options.horizontalCentered = bool(config.get("horizontalCenter", False))
        sheet.print_options.verticalCentered = bool(config.get("verticalCenter", False))
        sheet.print_options.gridLines = bool(config.get("showGridLines", False))
        sheet.print_title_rows = f"${min_row}:${min_row}" if config.get("repeatHeader") else None
        if config.get("freezeTopRow") and min_row < max_row:
            sheet.freeze_panes = sheet.cell(row=min_row + 1, column=min_col).coordinate
        else:
            sheet.freeze_panes = None
        set_margins(sheet, config.get("margin", "normal"))
        sheet.row_breaks.brk = []
        if int(config.get("fitToHeight", 0)) == 0:
            for row_start in range(min_row + preview_rows_per_page(config), max_row + 1, preview_rows_per_page(config)):
                sheet.row_breaks.append(Break(id=row_start))
        apply_wrapped_columns(sheet, min_row, min_col, max_row, max_col, config)
        if config.get("beautify", True):
            beautify_range(sheet, min_row, min_col, max_row, max_col)
        for column in range(min_col, max_col + 1):
            letter = get_column_letter(column)
            current = sheet.column_dimensions[letter]
            if not isinstance(current, ColumnDimension) or current.width is None:
                sheet.column_dimensions[letter].width = 14
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def pdf_data(formula_sheet, value_sheet, bounds: tuple[int, int, int, int]) -> list[list[str]]:
    min_row, min_col, max_row, max_col = bounds
    preview = page_cells(formula_sheet, value_sheet, min_row, max_row, min_col, max_col)
    if not preview:
        return [["Empty sheet"]]
    return [[str(value)[:48] for value in row] for row in preview]


def safe_pdf_name(name: str) -> str:
    return re.sub(r"[^\w\u4e00-\u9fff -]+", "_", name)


def export_print_pdf(
    source: str | Path,
    output: str | Path,
    selected_sheet_names: Iterable[Any],
    plan: str | dict[str, Any] | None = None,
) -> Path:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    source_path = Path(source)
    workbook_formula = load_workbook(source_path, data_only=False)
    workbook_values = load_workbook(source_path, data_only=True)
    targets = normalize_print_targets(selected_sheet_names)
    target_map = {target["sheet"]: target for target in targets}
    selected_names = set(target_map)
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    config = resolve_plan(plan or "fit-columns")
    page_size = landscape(A4) if config.get("orientation") == "landscape" else portrait(A4)
    document = SimpleDocTemplate(
        str(output_path),
        pagesize=page_size,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    styles["Title"].fontName = "STSong-Light"
    story = []
    for formula_sheet, value_sheet in zip(workbook_formula.worksheets, workbook_values.worksheets):
        if formula_sheet.title not in selected_names or formula_sheet.sheet_state != "visible":
            continue
        bounds, _ = target_bounds(formula_sheet, target_map[formula_sheet.title])
        story.append(Paragraph(f"{safe_pdf_name(source_path.stem)} - {safe_pdf_name(formula_sheet.title)}", styles["Title"]))
        data = pdf_data(formula_sheet, value_sheet, bounds)
        column_width = (page_size[0] - 16 * mm) / max(1, len(data[0]))
        table = Table(data, repeatRows=1, colWidths=[column_width] * len(data[0]))
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B7F72")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8D2CF")),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F3EE")]),
                ]
            )
        )
        story.append(table)
        story.append(PageBreak())
    if story and isinstance(story[-1], PageBreak):
        story.pop()
    if not story:
        story.append(Paragraph("No sheets selected.", styles["Title"]))
    document.build(story)
    return output_path


def export_macro_workbook(
    source: str | Path,
    output: str | Path,
    selected_sheet_names: Iterable[Any],
    plan: str | dict[str, Any],
) -> Path:
    targets = normalize_print_targets(selected_sheet_names)
    selected_names = [target["sheet"] for target in targets]
    if not selected_names:
        raise ValueError("No sheets selected for macro export")

    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    config = resolve_plan(plan)
    sheet_names_literal = ", ".join(f'"{name.replace("\"", "\"\"")}"' for name in selected_names)
    vba_code = f"""
Option Explicit

Public Sub PrintSelectedSheets()
    Dim targetSheets As Variant
    targetSheets = Array({sheet_names_literal})
    ThisWorkbook.Worksheets(targetSheets).PrintOut Copies:=1, Collate:=True
End Sub
""".strip()
    ps_code = f"""
$ErrorActionPreference = 'Stop'
$source = '{Path(source).resolve()}'
$output = '{output_path.resolve()}'
$sheetNames = @({", ".join("'" + name.replace("'", "''") + "'" for name in selected_names)})
$macroCode = @'
{vba_code}
'@
$excel = $null
$workbook = $null
try {{
  $excel = New-Object -ComObject Excel.Application
  $excel.Visible = $false
  $excel.DisplayAlerts = $false
  $workbook = $excel.Workbooks.Open($source)
  $summary = $workbook.Worksheets.Add($workbook.Worksheets.Item(1))
  $summary.Name = '批量打印台'
  $summary.Range('A1').Value2 = '启用宏后点击下方按钮打印本次选中的 Sheet'
  $summary.Range('A2').Value2 = '打印方案：{config.get("fitMode")}'
  $summary.Range('A3').Value2 = '纸张方向：{config.get("orientation")}'
  $summary.Range('A5').Value2 = '本次打印 Sheet'
  for ($i = 0; $i -lt $sheetNames.Count; $i++) {{
    $summary.Cells.Item(6 + $i, 1).Value2 = $sheetNames[$i]
  }}
  $button = $summary.Buttons().Add(24, 36, 180, 34)
  $button.Characters().Text = '一键打印所选 Sheet'
  $workbook.SaveAs($output, 52)
  $module = $workbook.VBProject.VBComponents.Add(1)
  $module.Name = 'BatchPrintModule'
  $module.CodeModule.AddFromString($macroCode)
  $summary = $workbook.Worksheets.Item('批量打印台')
  $summary.Buttons().Item(1).OnAction = 'PrintSelectedSheets'
  $workbook.Save()
}}
finally {{
  if ($workbook -ne $null) {{ $workbook.Close($false) }}
  if ($excel -ne $null) {{ $excel.Quit() }}
}}
""".strip()

    with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as script_file:
        script_file.write(ps_code)
        script_path = Path(script_file.name)

    try:
        subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
            ],
            check=True,
            capture_output=True,
            text=True,
        )
    except subprocess.CalledProcessError as error:
        stderr = (error.stderr or "").strip()
        stdout = (error.stdout or "").strip()
        stderr = stderr or stdout or "当前环境无法调用 Excel 桌面宏接口，暂时不能生成带一键打印的 VBA Excel 文件"
        raise RuntimeError(stderr) from error
    finally:
        script_path.unlink(missing_ok=True)

    return output_path


def dumps(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False)
